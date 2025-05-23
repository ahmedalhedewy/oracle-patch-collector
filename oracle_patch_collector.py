#!/usr/bin/env python3
"""
Oracle Patch Collector

This script connects to multiple Oracle servers via SSH,
retrieves patch information using OPatch commands from all Oracle homes,
and exports the results to an Excel file.

Author: Ahmed Alhedewy
Version: 1.0.0
License: MIT

Requirements:
- paramiko (for SSH connections)
- openpyxl (for Excel file creation)
- Install with: pip install -r requirements.txt
"""

import paramiko
import getpass
import time
import re
import os
from openpyxl import Workbook
from datetime import datetime

def connect_to_server(hostname, username, password, max_retries=1):
    """
    Establishes SSH connection to a server with retry capability
    """
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    
    retry_count = 0
    while retry_count <= max_retries:
        try:
            ssh.connect(hostname, username=username, password=password, timeout=10)
            return ssh
        except paramiko.AuthenticationException:
            print(f"Authentication failed for {hostname} with username {username}")
            return None
        except paramiko.SSHException as e:
            print(f"SSH error: {str(e)}")
            retry_count += 1
            if retry_count <= max_retries:
                print(f"Retrying connection ({retry_count}/{max_retries})...")
                time.sleep(2)
            else:
                print(f"Max retries reached for {hostname}")
                return None
        except Exception as e:
            print(f"Connection error to {hostname}: {str(e)}")
            return None

def run_command(ssh, command):
    """
    Executes a command on the server and returns the output
    """
    if ssh is None:
        return ""
    
    try:
        stdin, stdout, stderr = ssh.exec_command(command)
        output = stdout.read().decode('utf-8')
        error = stderr.read().decode('utf-8')
        
        if error and not "ORA-" in error:  # Ignore common Oracle errors
            print(f"Error executing command: {error}")
        
        return output
    except Exception as e:
        print(f"Error running command: {str(e)}")
        return ""

def find_oracle_homes(ssh):
    """
    Finds all Oracle home directories on the server
    """
    if ssh is None:
        return []
    
    # Various commands to detect Oracle homes
    # First try the oratab file which is the most reliable source
    oratab_cmd = "cat /etc/oratab 2>/dev/null || cat /var/opt/oracle/oratab 2>/dev/null"
    oratab_output = run_command(ssh, oratab_cmd)
    
    oracle_homes = []
    
    # Parse oratab output to extract Oracle homes
    if oratab_output:
        for line in oratab_output.splitlines():
            # Skip comments and empty lines
            if line.startswith('#') or not line.strip():
                continue
            
            # Format is usually: SID:ORACLE_HOME:START
            parts = line.split(':')
            if len(parts) >= 2 and parts[1] and parts[1] != '/':
                oracle_homes.append(parts[1].strip())
    
    # If we didn't find any homes in oratab, try other methods
    if not oracle_homes:
        # Try to find Oracle homes by searching for common Oracle directories
        find_cmd = "find /u01 /opt /oracle -name dbhome_1 -o -name dbhome_2 -o -name dbhome_3 2>/dev/null"
        find_output = run_command(ssh, find_cmd)
        
        if find_output:
            for line in find_output.splitlines():
                if line.strip():
                    # The parent directory of dbhome_N is typically an Oracle home
                    oracle_home = os.path.dirname(line.strip())
                    oracle_homes.append(oracle_home)
    
    # If we still didn't find any homes, try checking environment variables
    if not oracle_homes:
        env_cmd = "env | grep ORACLE_HOME"
        env_output = run_command(ssh, env_cmd)
        
        if env_output:
            for line in env_output.splitlines():
                if "ORACLE_HOME=" in line:
                    oracle_home = line.split("=")[1].strip()
                    oracle_homes.append(oracle_home)
    
    # Remove duplicates while preserving order
    unique_homes = []
    for home in oracle_homes:
        if home not in unique_homes:
            unique_homes.append(home)
    
    # Add a default location as fallback if nothing was found
    if not unique_homes:
        unique_homes.append("/u01/app/oracle/product/19.3.0.0/dbhome_1")
        print("No Oracle homes found. Using default fallback location.")
    
    return unique_homes

def extract_release_info(description):
    """
    Attempts to extract release information from patch description
    """
    # Look for patterns like "Release x.x.x.x" or "version x.x.x.x"
    release_match = re.search(r'[Rr]elease\s+([\d\.]+)', description)
    if release_match:
        return release_match.group(1)
    
    version_match = re.search(r'[Vv]ersion\s+([\d\.]+)', description)
    if version_match:
        return version_match.group(1)
    
    # Look for any version-like pattern in the description
    generic_version = re.search(r'(\d+\.\d+\.\d+\.\d+)', description)
    if generic_version:
        return generic_version.group(1)
    
    return ""

def version_is_newer(ver1, ver2):
    """
    Compares two version strings and returns True if ver1 is newer than ver2
    """
    # Split version strings into components
    ver1_parts = [int(x) for x in ver1.split('.') if x.isdigit()]
    ver2_parts = [int(x) for x in ver2.split('.') if x.isdigit()]
    
    # Make sure both lists have the same length
    while len(ver1_parts) < len(ver2_parts):
        ver1_parts.append(0)
    while len(ver2_parts) < len(ver1_parts):
        ver2_parts.append(0)
    
    # Compare part by part
    for i in range(len(ver1_parts)):
        if ver1_parts[i] > ver2_parts[i]:
            return True
        elif ver1_parts[i] < ver2_parts[i]:
            return False
    
    # If we get here, versions are equal
    return False

def get_oracle_patches(ssh):
    """
    Retrieves Oracle patch information from all Oracle homes
    """
    if ssh is None:
        return []
    
    # Find all Oracle homes
    oracle_homes = find_oracle_homes(ssh)
    patch_results = []
    
    for oracle_home in oracle_homes:
        print(f"Checking Oracle home: {oracle_home}")
        
        # Check if OPatch exists in this home
        opatch_check_cmd = f"ls -l {oracle_home}/OPatch/opatch 2>/dev/null"
        opatch_check_result = run_command(ssh, opatch_check_cmd)
        
        if not opatch_check_result or "No such file" in opatch_check_result:
            print(f"  OPatch not found in {oracle_home}, skipping...")
            continue
        
        # Run OPatch commands
        lspatches_cmd = f"{oracle_home}/OPatch/opatch lspatches"
        version_cmd = f"{oracle_home}/OPatch/opatch version"
        
        lspatches_output = run_command(ssh, lspatches_cmd)
        version_output = run_command(ssh, version_cmd)
        
        # Also get Oracle database version for this home
        oracle_version_cmd = f"{oracle_home}/bin/sqlplus -V 2>/dev/null"
        oracle_version_output = run_command(ssh, oracle_version_cmd)
        
        # Get SID/database name associated with this ORACLE_HOME if possible
        sid_cmd = f"ps -ef | grep pmon | grep {oracle_home} | awk '{{print $NF}}' | sed 's/ora_pmon_//g'"
        sid_output = run_command(ssh, sid_cmd).strip()
        
        # If we couldn't get SID from processes, try to derive it from the path
        if not sid_output:
            # Try to extract SID from path (common naming convention)
            path_parts = oracle_home.split('/')
            for part in path_parts:
                if part.startswith('db_') or part.startswith('ora'):
                    sid_output = part
                    break
        
        patch_results.append({
            "oracle_home": oracle_home,
            "sid": sid_output,
            "oracle_version": oracle_version_output,
            "lspatches": lspatches_output,
            "version": version_output
        })
    
    return patch_results

def parse_patch_info(patch_info):
    """
    Parses OPatch output to extract patch information
    """
    results = {
        "oracle_home": patch_info["oracle_home"],
        "sid": patch_info["sid"],
        "database_release": "",
        "ojvm_release": "",
        "ocw_release": "",
        "opatch_version": "",
        "oracle_version": ""
    }
    
    # Extract Oracle version information if available
    oracle_version_match = re.search(r'Version (\d+\.\d+\.\d+\.\d+)', patch_info["oracle_version"])
    if oracle_version_match:
        results["oracle_version"] = oracle_version_match.group(1)
    
    # Parse lspatches output
    for line in patch_info["lspatches"].split('\n'):
        match = re.search(r'(\d+);(.*)', line)
        if match:
            description = match.group(2).strip()
            
            # Determine patch type and release information
            description_lower = description.lower()
            
            # Check if it's a database patch
            if "database" in description_lower or "db" in description_lower:
                db_release = extract_release_info(description)
                if db_release and (not results["database_release"] or version_is_newer(db_release, results["database_release"])):
                    results["database_release"] = db_release
            
            # Check if it's an OJVM patch
            elif "ojvm" in description_lower or "java" in description_lower:
                ojvm_release = extract_release_info(description)
                if ojvm_release and (not results["ojvm_release"] or version_is_newer(ojvm_release, results["ojvm_release"])):
                    results["ojvm_release"] = ojvm_release
            
            # Check if it's an OCW patch
            elif "ocw" in description_lower or "client" in description_lower:
                ocw_release = extract_release_info(description)
                if ocw_release and (not results["ocw_release"] or version_is_newer(ocw_release, results["ocw_release"])):
                    results["ocw_release"] = ocw_release
    
    # Parse version output
    version_match = re.search(r'Version: ([\d\.]+)', patch_info["version"])
    if version_match:
        results["opatch_version"] = version_match.group(1)
    
    return results

def export_to_excel(server_data, filename):
    """
    Exports the collected data to an Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Oracle Patch Info"
    
    # Write headers
    headers = ["Hostname", "SID", "Oracle Home", "Oracle Version", "OPatch Version", 
               "Database Release", "OJVM RELEASE", "OCW RELEASE"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data
    row = 2
    for server, homes_data in server_data.items():
        for data in homes_data:
            ws.cell(row=row, column=1, value=server)
            ws.cell(row=row, column=2, value=data["sid"])
            ws.cell(row=row, column=3, value=data["oracle_home"])
            ws.cell(row=row, column=4, value=data["oracle_version"])
            ws.cell(row=row, column=5, value=data["opatch_version"])
            ws.cell(row=row, column=6, value=data["database_release"])
            ws.cell(row=row, column=7, value=data["ojvm_release"])
            ws.cell(row=row, column=8, value=data["ocw_release"])
            row += 1
    
    # Save the workbook
    wb.save(filename)
    print(f"Data exported to {filename}")

def main():
    # Get list of servers
    servers_input = input("Enter IP addresses/hostnames (comma-separated or from a file path): ")
    
    if os.path.isfile(servers_input):
        with open(servers_input, 'r') as f:
            servers = [line.strip() for line in f if line.strip()]
    else:
        servers = [s.strip() for s in servers_input.split(',')]
    
    # Try with default Oracle username
    username = "oracle"
    print(f"Attempting to connect with default username: {username}")
    password = getpass.getpass(f"Enter password for {username}: ")
    
    # Prepare output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"oracle_patches_{timestamp}.xlsx"
    
    # Process each server
    server_data = {}
    
    for server in servers:
        print(f"\nConnecting to {server}...")
        ssh = connect_to_server(server, username, password)
        
        # If connection fails, try with different credentials
        if ssh is None:
            print(f"Failed to connect to {server} with default credentials.")
            retry = True
            while retry:
                retry_option = input("Would you like to try with different credentials? (y/n): ")
                if retry_option.lower() == 'y':
                    username = input("Enter username: ")
                    password = getpass.getpass(f"Enter password for {username}: ")
                    ssh = connect_to_server(server, username, password)
                    if ssh is None:
                        print(f"Failed to connect to {server}. Authentication failed.")
                    else:
                        retry = False  # Successfully connected
                else:
                    print(f"Skipping {server}.")
                    retry = False  # User chose not to retry
        
        if ssh:
            print(f"Getting patch information from {server}...")
            patch_info_list = get_oracle_patches(ssh)
            
            if patch_info_list:
                server_data[server] = []
                for patch_info in patch_info_list:
                    parsed_info = parse_patch_info(patch_info)
                    server_data[server].append(parsed_info)
                print(f"Found {len(patch_info_list)} Oracle homes on {server}")
            else:
                print(f"No Oracle installations found on {server}")
                server_data[server] = []
            
            # Close connection
            ssh.close()
            print(f"Completed gathering information from {server}")
        
        # Small delay to prevent overwhelming connections
        time.sleep(1)
    
    # Export data to Excel if we collected any data
    if server_data:
        export_to_excel(server_data, output_file)
    else:
        print("No data collected. Excel file not created.")

if __name__ == "__main__":
    main()
